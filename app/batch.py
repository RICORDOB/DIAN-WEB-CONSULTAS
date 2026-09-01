"""
batch — Consultas masivas (panel Contadores).

Procesa un archivo .xlsx con las columnas:
  - tipo_documento   (obligatoria)
  - numero_documento (obligatoria)
  - contrasena       (obligatoria)
  - fecha_vencimiento (opcional; si falta se calcula del calendario)
  - estado           (opcional; filas con 'ok' se saltan en re-ejecución)

Cada fila se procesa reutilizando DianRunner.consulta_individual(), que ya
aisla cada consulta en su propio directorio de trabajo. Al terminar se actualiza
la columna 'estado' del Excel de entrada y se recogen todos los .xls generados.
"""

from __future__ import annotations

import asyncio
from pathlib import Path

from openpyxl import Workbook, load_workbook

from .runner import DianRunner

# Valores de la columna 'estado'
OK = "ok"
ERROR_CREDENCIALES = "error_credenciales"
DESCONOCIDO = "desconocido"
EXCEPCION = "excepcion"

COLUMNAS_REQUERIDAS = ("tipo_documento", "numero_documento", "contrasena")


def _normalizar(texto: str | None) -> str:
    return (texto or "").strip().lower()


def cargar_filas(ruta: Path) -> tuple[list[dict], list[str]]:
    """Lee el Excel de entrada y devuelve (filas, encabezados).

    Las filas con tipo/numero/contrasena vacíos se omiten; las que ya tienen
    estado 'ok' se omiten para permitir re-ejecuciones resumibles.
    """
    wb = load_workbook(ruta)
    ws = wb.active
    filas_raw = list(ws.iter_rows(values_only=True))
    if not filas_raw:
        return [], []

    encabezados = [_normalizar(c) for c in filas_raw[0]]
    idx = {nombre: i for i, nombre in enumerate(encabezados)}
    faltantes = [c for c in COLUMNAS_REQUERIDAS if c not in idx]
    if faltantes:
        raise ValueError(
            "El archivo debe tener las columnas: "
            + ", ".join(COLUMNAS_REQUERIDAS)
            + f". Faltan: {', '.join(faltantes)}."
        )

    filas: list[dict] = []
    for i, row in enumerate(filas_raw[1:], start=2):  # i = fila Excel (1-indexed)
        tipo = row[idx["tipo_documento"]]
        numero = row[idx["numero_documento"]]
        password = row[idx["contrasena"]]
        if tipo is None and numero is None and password is None:
            continue
        if numero is None or str(numero).strip() == "" or password is None:
            continue
        if _normalizar(row[idx.get("estado", 0)]) == OK:
            continue
        filas.append({
            "fila_excel": i,
            "tipo_documento": str(tipo or "Cédula de Ciudadanía").strip(),
            "numero_documento": str(numero).strip(),
            "contrasena": str(password),
        })
    return filas, encabezados


def guardar_estado(ruta: Path, fila_excel: int, estado: str) -> None:
    """Escribe el resultado en la columna 'estado' de la fila dada."""
    _celda_columna(ruta, COLUMNA_ESTADO, fila_excel, estado)


# Columnas que alimenta el sistema al terminar el proceso de cada cliente
COLUMNA_ESTADO = "estado"
COLUMNA_FECHA_VENCIMIENTO = "fecha_vencimiento"

PLANTILLA_ENCABEZADOS = (
    "tipo_documento",
    "numero_documento",
    "contrasena",
    COLUMNA_FECHA_VENCIMIENTO,
    COLUMNA_ESTADO,
)


def generar_plantilla(ruta: Path) -> None:
    """Crea una plantilla .xlsx con las columnas esperadas, sin fila de ejemplo.

    La primera hoja es la activa (cargar_filas/guardar_* usan ws.active).
    'fecha_vencimiento' y 'estado' las alimenta el sistema al terminar cada
    cliente; no deben incluirse datos ficticios que se procesarían como reales.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "clientes"
    for col, nombre in enumerate(PLANTILLA_ENCABEZADOS, start=1):
        ws.cell(row=1, column=col, value=nombre)
    wb.save(ruta)


def _celda_columna(ruta: Path, nombre: str, fila_excel: int, valor) -> None:
    """Escribe 'valor' en la columna de encabezado 'nombre' de la fila dada.

    Crea la columna al final si aún no existe (mismo criterio que guardar_estado)
    para que el proceso pueda alimentar el resultado aunque el .xlsx del
    contador tuviera otras columnas.
    """
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [_normalizar(c) for c in next(ws.iter_rows(values_only=True))]
    nombre = _normalizar(nombre)
    if nombre not in encabezados:
        ws.cell(row=1, column=len(encabezados) + 1, value=nombre)
        encabezados.append(nombre)
    col = encabezados.index(nombre) + 1
    ws.cell(row=fila_excel, column=col, value=valor)
    wb.save(ruta)


def guardar_fecha_vencimiento(ruta: Path, fila_excel: int, fecha: str) -> None:
    """Escribe la fecha de vencimiento calculada en la fila dada."""
    _celda_columna(ruta, COLUMNA_FECHA_VENCIMIENTO, fila_excel, fecha)


def _estado_final(exc: Exception) -> str:
    msg = str(exc).lower()
    if "credenciales" in msg or "contraseña" in msg or "clave" in msg:
        return ERROR_CREDENCIALES
    if "determinar" in msg:
        return DESCONOCIDO
    return EXCEPCION


async def ejecutar_batch(
    job_dir: Path,
    entrada: Path,
    progreso: callable,
) -> dict:
    """Procesa todas las filas del Excel de forma secuencial.

    progreso(grupo, mensaje, detalle) permite al orquestador reportar avance.
    Devuelve un resumen con {total, ok, error_credenciales, desconocido,
    excepcion, generados: [paths]}.
    """
    filas, _ = cargar_filas(entrada)
    if not filas:
        progreso("info", "El archivo no tiene filas por procesar.", "")
        return {"total": 0, "ok": 0, "generados": [], "detalle": []}

    resumen = {"total": len(filas), "ok": 0, "error_credenciales": 0,
               "desconocido": 0, "excepcion": 0, "generados": [], "detalle": []}

    # Reutilizamos UN solo DianRunner cuyo job_dir aloja todos los .xls de la
    # carpeta 'clientes'. consulta_individual() lanza su propio navegador por
    # llamada; esto mantiene el aislamiento por fila y simplifica el progreso.
    for i, creds in enumerate(filas, start=1):
        numero = creds["numero_documento"]
        progreso("fila", f"({i}/{len(filas)}) Procesando {creds['tipo_documento']} {numero}...", creds)

        runner = DianRunner(job_dir=job_dir)
        fila_resultado = {"fila_excel": creds["fila_excel"], "numero_documento": numero}
        estado = OK
        try:
            final = await runner.consulta_individual(
                creds["tipo_documento"], numero, creds["contrasena"]
            )
            resumen["ok"] += 1
            resumen["generados"].append(str(final))
            final_name = Path(final).name
            guardar_estado(entrada, creds["fila_excel"], OK)
            # Alimenta la fecha de vencimiento que calculó el runner (calendario DIAN)
            if runner.ultima_fecha_vencimiento:
                guardar_fecha_vencimiento(
                    entrada, creds["fila_excel"], runner.ultima_fecha_vencimiento
                )
            fila_resultado["final"] = final_name
        except Exception as exc:  # noqa: BLE001
            estado = _estado_final(exc)
            resumen[estado] += 1
            guardar_estado(entrada, creds["fila_excel"], estado)
            fila_resultado["error"] = str(exc)

        resumen["detalle"].append(fila_resultado)

    progreso("resumen",
             f"Proceso finalizado: OK={resumen['ok']} "
             f"error_credenciales={resumen['error_credenciales']} "
             f"desconocido={resumen['desconocido']} "
             f"excepcion={resumen['excepcion']} de {resumen['total']}.",
             "")
    return resumen
