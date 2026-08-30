"""
DianRunner — Núcleo de consulta DIAN (login + exógena + renta + FE).

Refactor de dian_login.py para funcionar como un servicio web: en vez de leer
credenciales desde un Excel local y orquestar desde consola, esta clase recibe
los datos de UNA consulta individual por parámetro y produce el MISMO libro
.xls de 3 hojas que generaba el script original.

Diseño:
  - Cada ejecución se aísla en un directorio de trabajo (job_dir) propio, de
    modo que varias solicitudes no compartan descargas ni salidas.
  - UN solo cliente por consulta (flujo individual). El procesamiento masivo
    futuro reutilizará la misma clase por fila.
  - El navegador (Playwright/Chromium) se lanza en modo headless con flags
    compatibles con contenedores de baja RAM (--no-sandbox, --disable-dev-shm).
  - El progreso se reporta mediante un callback opcional (para el frontend).
"""

from __future__ import annotations

import asyncio
import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright.async_api import async_playwright

from .comun import (
    ANIO_EXOGENO,
    AVISO_INFO_EXOGENA,
    CALENDARIO_RENTA_2026,
    NORMA_TOPES,
    REINTENTOS,
    RESOLUCION_UVT,
    TEXTO_SIN_DATOS_FE,
    TIMEOUT_DESCARGA,
    TOPES,
    URL_LOGIN,
    UVT_2025,
)

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------

# Fase 4 desactivada: los resultados se quedan en el entorno de ejecución
# (la subida a Google Drive puede reactivarse en una versión futura).
SUBIR_A_DRIVE = False

# Texto de la notificación de progreso en login (único de este módulo)
MENSAJE_LOGIN = "Ingresando al portal DIAN"

ProgresoCb = Callable[[str, bool], None]


def num_documento_mascarado(numero_documento: str) -> str:
    """Máscara simple para no imprimir el documento completo en los logs."""
    n = numero_documento.strip()
    if len(n) <= 4:
        return "*" * len(n)
    return n[:2] + "*" * (len(n) - 4) + n[-2:]


class DianRunner:
    """Ejecuta una consulta individual contra el portal DIAN y genera el libro."""

    def __init__(self, job_dir: Path, progreso: Optional[ProgresoCb] = None) -> None:
        self.job_dir = job_dir
        self.download_dir = job_dir / "descargas"
        self.clientes_dir = job_dir / "clientes"
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self.clientes_dir.mkdir(parents=True, exist_ok=True)
        self.log_path = self.download_dir / f"log_dian_{datetime.now():%Y%m%d_%H%M%S}.txt"
        self.progreso = progreso
        # Análisis de renta de la última consulta, para que el orquestador
        # (main.py) pueda exponerlo al frontend vía /api/job/{id}.
        self.ultimo_analisis: dict | None = None

    # -------------------------------------------------------------------
    # Logging / progreso
    # -------------------------------------------------------------------
    def loguear(self, mensaje: str) -> None:
        """Escribe a consola, al archivo de log del job y, si hay callback, al frontend."""
        ts = datetime.now().strftime("%H:%M:%S")
        linea = f"[{ts}] {mensaje}"
        print(linea)
        with open(self.log_path, "a", encoding="utf-8") as fh:
            fh.write(linea + "\n")
        if self.progreso:
            self.progreso(linea, False)

    def emitir_done(self, mensaje: str = "") -> None:
        if self.progreso:
            self.progreso(mensaje or "Proceso finalizado", True)

    # -------------------------------------------------------------------
    # Helpers
    # -------------------------------------------------------------------
    @staticmethod
    def _fecha_vencimiento_renta(numero_documento: str) -> date:
        """Vencimiento de renta AG 2025 según los dos últimos dígitos del NIT."""
        ultimos_dos = int(numero_documento.strip()[-2:])
        for (a, b), fecha in CALENDARIO_RENTA_2026.items():
            if ultimos_dos == a or ultimos_dos == b:
                return fecha
        raise ValueError(
            f"No hay fecha en el calendario para los dígitos '{ultimos_dos:02d}'"
        )

    @staticmethod
    def _nombre_seguro(nombre: str) -> str:
        """Elimina caracteres no válidos en nombres de archivo."""
        return re.sub(r'[\\/*?:<>|"]', "", nombre).strip()

    async def _cerrar_modales_dian(self, page) -> None:
        """Cierra modales visibles de la DIAN (contraseña por vencer/vencida)."""
        for _ in range(40):
            intentado = False
            for texto in ("Cerrar", "Aceptar"):
                try:
                    btn = page.locator(f"text={texto}").first
                    if await btn.is_visible(timeout=500):
                        await btn.click()
                        await page.wait_for_timeout(500)
                        intentado = True
                        break
                except Exception:
                    pass
            if not intentado:
                return

    # -------------------------------------------------------------------
    # Fase 1: Login
    # -------------------------------------------------------------------
    async def iniciar_sesion(self, page, creds: dict) -> str:
        """Login en MUISCA. Devuelve 'ok', 'error_credenciales' o 'desconocido'."""
        await page.goto(URL_LOGIN, wait_until="networkidle")
        await page.get_by_role("button", name="A nombre propio", exact=True).click()
        await page.locator("mat-select[name='tipoDocumento']").click()
        await page.get_by_role(
            "option", name=creds["tipo_documento"], exact=False
        ).first.click()
        await page.fill("input[name='numDocumento']", creds["numero_documento"])
        await page.fill("input[name='password']", creds["contrasena"])
        await page.locator("mat-checkbox[name='aceptaTratamientoDatos']").click()
        await page.get_by_role("button", name="Ingresar", exact=True).click()

        # Si el formulario desapareció, entró
        try:
            await page.wait_for_selector(
                "mat-select[name='tipoDocumento']", state="detached", timeout=10000
            )
            return "ok"
        except Exception:
            pass

        error_loc = page.locator("text=/inválid|incorrect|no coincid|error/i")
        if await error_loc.count():
            return "error_credenciales"
        return "desconocido"

    async def _intentar_login(self, page, creds: dict) -> str:
        """Reintenta solo errores transitorios; no reintenta credenciales rechazadas."""
        ultimo = None
        for intento in range(1, REINTENTOS + 1):
            try:
                return await self.iniciar_sesion(page, creds)
            except Exception as exc:
                ultimo = exc
                self.loguear(f"  [login][reintento {intento}/{REINTENTOS}] "
                             f"{type(exc).__name__}: {exc}")
                try:
                    await page.goto(URL_LOGIN, wait_until="networkidle")
                except Exception:
                    pass
                await asyncio.sleep(1)
        raise ultimo

    # -------------------------------------------------------------------
    # Fase 2 y 3: descargas y análisis
    # -------------------------------------------------------------------
    async def consultar_exogena(self, page, anio: str, nro_doc: str) -> Path:
        """Descarga el reporte de información exógena en Excel."""
        await page.wait_for_timeout(1000)
        await self._cerrar_modales_dian(page)
        await page.locator("input[id='vistaDashboard:frmDashboard:btnExogena']").click(force=True)
        await page.wait_for_selector(
            "input[id='vistaDashboard:frmDashboard:btnBuscar']",
            state="visible", timeout=15000,
        )
        await self._cerrar_modales_dian(page)
        await page.locator("input[id='vistaDashboard:frmDashboard:btnBuscar']").click(force=True)
        await page.wait_for_selector(
            "select[id='vistaDashboard:frmDashboard:anioSel']",
            state="visible", timeout=15000,
        )
        await page.select_option(
            "select[id='vistaDashboard:frmDashboard:anioSel']", value=anio
        )
        await page.wait_for_timeout(2000)

        destino = self.download_dir / f"informacion_exogena_{anio}_{nro_doc}.xls"
        for intento in range(1, REINTENTOS + 1):
            try:
                await self._cerrar_modales_dian(page)
                async with page.expect_download(timeout=TIMEOUT_DESCARGA) as dl_info:
                    await page.locator(
                        "input[id='vistaDashboard:frmDashboard:btnExogenaGenerar']"
                    ).click(force=True)
                    await self._cerrar_modales_dian(page)
                    download = await dl_info.value
                await download.save_as(destino)
                if destino.exists() and destino.stat().st_size > 0:
                    break
                raise RuntimeError("la descarga quedó vacía")
            except Exception as exc:
                self.loguear(f"  [exogena][reintento {intento}/{REINTENTOS}] "
                             f"{type(exc).__name__}: {exc}")
                if intento == REINTENTOS:
                    raise
                await page.wait_for_timeout(1500)
        return destino

    async def consultar_facturacion_electronica(
        self, page, anio: str, nro_doc: str
    ) -> Optional[Path]:
        """Descarga el reporte de facturación electrónica. Devuelve None si no hay datos."""
        try:
            await self._cerrar_modales_dian(page)
            await page.locator("input[id='vistaDashboard:frmDashboard:btnFE']").click(force=True)
            await page.wait_for_selector(
                "input[id='vistaDashboard:frmDashboard:btnBuscarFE']",
                state="visible", timeout=15000,
            )
            await self._cerrar_modales_dian(page)
            await page.locator("input[id='vistaDashboard:frmDashboard:btnBuscarFE']").click(force=True)
            await page.wait_for_selector(
                "select[id='vistaDashboard:frmDashboard:anioSelFE']",
                state="visible", timeout=15000,
            )
            await page.select_option(
                "select[id='vistaDashboard:frmDashboard:anioSelFE']", value=anio
            )
            await page.wait_for_timeout(2000)

            destino = self.download_dir / f"facturacion_{anio}_{nro_doc}.xls"
            for intento in range(1, REINTENTOS + 1):
                try:
                    await self._cerrar_modales_dian(page)
                    async with page.expect_download(timeout=TIMEOUT_DESCARGA) as dl_info:
                        await page.locator(
                            "input[id='vistaDashboard:frmDashboard:btnFFGenerar']"
                        ).click(force=True)
                        await self._cerrar_modales_dian(page)
                        download = await dl_info.value
                    await download.save_as(destino)
                    if destino.exists() and destino.stat().st_size > 0:
                        return destino
                    raise RuntimeError("la descarga quedó vacía")
                except Exception as exc:
                    if await page.get_by_text(
                        TEXTO_SIN_DATOS_FE, exact=False
                    ).count():
                        self.loguear("  [FE] sin información para el año seleccionado "
                                     "-> se omite hoja 3.")
                        return None
                    self.loguear(f"  [FE][reintento {intento}/{REINTENTOS}] "
                                 f"{type(exc).__name__}: {exc}")
                    if intento == REINTENTOS:
                        break
                    await page.wait_for_timeout(1500)
        except Exception as exc:
            self.loguear(f"  [FE][aviso] No se pudo descargar la facturación electrónica "
                         f"({type(exc).__name__}: {exc}) -> se omite hoja 3.")
            return None
        return None

    @staticmethod
    def _copiar_valores(origen_ws, destino_ws) -> None:
        for row in origen_ws.iter_rows():
            for cell in row:
                destino_ws[cell.coordinate].value = cell.value

    def armar_libro_cliente(self, exogena_path: Path, analisis: dict,
                            facturacion_path: Optional[Path]) -> Path:
        """Copia el XLSX de exógena y le agrega las hojas Renta y Facturación Electrónica."""
        nombre = self._nombre_seguro(analisis["nombre_cliente"])
        tmp_exo = self.download_dir / f".tmp_exo_{nombre}.xlsx"
        shutil.copy(exogena_path, tmp_exo)
        wb = load_workbook(str(tmp_exo))
        wb.worksheets[0].title = "Información Exógena"

        ws2 = wb.create_sheet("Renta")
        ws2["A1"] = "Nombre del cliente"
        ws2["B1"] = analisis["nombre_cliente"]
        ws2["A2"] = "¿Declara renta?"
        ws2["B2"] = analisis["declara_renta"]
        ws2["A3"] = "Vence (Renta AG 2025)"
        ws2["B3"] = analisis.get("fecha_vencimiento", "")
        ws2["A5"] = "Razones:"
        ws2["A6"] = analisis["razones"]
        ws2["A6"].alignment = Alignment(wrap_text=True, vertical="top")

        ws3 = wb.create_sheet("Facturación Electrónica")
        if facturacion_path and facturacion_path.exists():
            tmp_fe = self.download_dir / f".tmp_fe_{nombre}.xlsx"
            shutil.copy(facturacion_path, tmp_fe)
            self._copiar_valores(load_workbook(str(tmp_fe)).active, ws3)
            tmp_fe.unlink()
        else:
            ws3["A1"] = TEXTO_SIN_DATOS_FE + "."

        final = self.clientes_dir / f"{nombre}.xls"
        wb.save(str(final))
        tmp_exo.unlink()
        return final

    def analizar_exogena(self, ruta: Path) -> dict:
        """Determina si la persona está obligada a declarar renta según los topes."""
        try:
            df = pd.read_excel(ruta, header=None, engine="openpyxl")
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(
                f"El reporte de exógena descargado es ilegible o está corrupto "
                f"({type(exc).__name__}: {exc})."
            ) from exc

        nombre = None
        for r in range(df.shape[0]):
            for c in range(df.shape[1]):
                cell = df.iloc[r, c]
                if isinstance(cell, str) and "Nombres / Razón social" in cell:
                    for cc in range(df.shape[1]):
                        val = df.iloc[r, cc]
                        if isinstance(val, str) and "Nombres" not in val and val.strip():
                            nombre = val.strip()
                            break
                    if nombre:
                        break
            if nombre:
                break

        valores = {}
        for r in range(df.shape[0]):
            if df.shape[1] <= 4:
                break
            v0 = df.iloc[r, 4]
            if isinstance(v0, str) and v0.strip().lower().startswith("tope"):
                label = v0.strip()
                raw = df.iloc[r, 5] if df.shape[1] > 5 else None
                try:
                    num = float(raw)
                except (TypeError, ValueError):
                    num = 0.0
                cat = None
                for key in ("Ingresos", "Patrimonio", "Consumo TC",
                            "Movimiento", "Compras"):
                    if key.lower() in label.lower():
                        cat = key
                        break
                if cat is None:
                    cat = label
                valores[cat] = num

        # Topes evaluados en forma estructurada (para mostrar en el panel)
        # y en forma de texto legible (para el libro Excel / razones).
        topes = []
        lineas = []
        declara = False
        for cat, uvt, desc, op in TOPES:
            valor = valores.get(cat, 0.0)
            umbral = uvt * UVT_2025
            excede = (valor >= umbral) if op == ">=" else (valor > umbral)
            if excede:
                declara = True
            topes.append({
                "desc": desc,
                "cat": cat,
                "reportado": valor,
                "umbral": umbral,
                "excede": excede,
            })
            lineas.append(
                f"- {desc} (Tope {cat}): reportado ${valor:,.0f}  |  "
                f"umbral {uvt:,} UVT = ${umbral:,.0f}  ->  "
                f"{'EXCEDE' if excede else 'no excede'}"
            )

        cabecera = (
            "La persona ESTÁ OBLIGADA a presentar declaración de renta (AG 2025) "
            "porque supera al menos uno de los topes del " + NORMA_TOPES + "."
            if declara else
            "La persona NO está obligada a declarar renta (AG 2025) según los "
            "topes del " + NORMA_TOPES + " (no supera ninguno)."
        )
        cuerpo = (
            f"Cotejo con UVT 2025 = ${UVT_2025:,} ({RESOLUCION_UVT}).\n"
            + "\n".join(lineas)
        )
        nota = (
            "\n\nNota: el reporte de exógena de la DIAN incluye consumos con "
            "tarjeta débito en 'Consumo TC'; la norma considera solo tarjeta de "
            "crédito, por lo que ese tope podría estar sobreestimado."
            if "Consumo TC" in valores else ""
        )
        razones = cabecera + "\n" + cuerpo + nota + "\n\n" + AVISO_INFO_EXOGENA

        return {
            "nombre_cliente": nombre or "DESCONOCIDO",
            "declara_renta": "Sí" if declara else "No",
            "razones": razones,
            "cabecera": cabecera,
            "nota": nota.strip(),
            "topes": topes,
        }

    # -------------------------------------------------------------------
    # Orquestación: consulta individual
    # -------------------------------------------------------------------
    async def _post_login(self, page, creds: dict) -> Path:
        """Tras el login: descarga exógena + FE, analiza renta y arma el libro."""
        await page.wait_for_timeout(2000)
        await self._cerrar_modales_dian(page)
        self.loguear("  [Fase2] Consultando información exógena...")
        exogena = await self.consultar_exogena(page, ANIO_EXOGENO, creds["numero_documento"])
        self.loguear(f"  [Fase2] Exógena descargada: {exogena}")

        self.loguear("  [Fase3] Analizando obligación de declarar renta...")
        analisis = self.analizar_exogena(exogena)
        analisis["fecha_vencimiento"] = creds.get("fecha_vencimiento", "")
        self.ultimo_analisis = analisis
        self.loguear(f"  [Fase3] {analisis['nombre_cliente']} -> declara renta: "
                     f"{analisis['declara_renta']} | vence: {analisis['fecha_vencimiento']}")

        self.loguear("  [Fase2] Consultando facturación electrónica...")
        facturacion = await self.consultar_facturacion_electronica(
            page, ANIO_EXOGENO, creds["numero_documento"]
        )

        final = self.armar_libro_cliente(exogena, analisis, facturacion)
        self.loguear(f"  [ok] Archivo cliente generado: {final}")
        return final

    async def consulta_individual(
        self,
        tipo_documento: str,
        numero_documento: str,
        contrasena: str,
    ) -> Path:
        """
        Ejecuta la consulta completa de UN cliente y devuelve la ruta del libro .xls.

        Lanza excepciones con mensajes legibles si el login falla o si el proceso
        no puede completarse. La credencial nunca se persiste: existe solo en memoria.
        """
        creds = {
            "tipo_documento": tipo_documento,
            "numero_documento": numero_documento,
            "contrasena": contrasena,
            "fecha_vencimiento": self._fecha_vencimiento_renta(numero_documento).isoformat(),
        }
        self.loguear(
            f"[info] Consulta individual: {tipo_documento} {numero_documento}"
        )
        self.emitir_done(
            f"Ingresando al portal DIAN ({num_documento_mascarado(numero_documento)})..."
        )

        async with async_playwright() as p:
            browser = await p.chromium.launch(
                headless=True,
                args=[
                    "--no-sandbox",
                    "--disable-dev-shm-usage",
                    "--disable-gpu",
                    "--single-process",
                ],
            )
            try:
                context = await browser.new_context(accept_downloads=True, locale="es-CO")
                page = await context.new_page()

                self.loguear(f"  [Paso 1/4] Login en MUISCA...")
                estado_login = await self._intentar_login(page, creds)
                if estado_login != "ok":
                    detalle = {
                        "error_credenciales": "Credenciales rechazadas o mensaje de error en la página.",
                        "desconocido": "No se pudo determinar el resultado del login.",
                    }[estado_login]
                    await page.screenshot(path=self.download_dir / "error_login.png")
                    raise RuntimeError(detalle)

                self.loguear("  [ok] Sesión iniciada correctamente.")
                self.emitir_done("Sesión iniciada. Descargando reportes...")
                final = await self._post_login(page, creds)

                self.emitir_done("Consulta completada. Libro generado.")
                return final
            finally:
                await context.close()
                await browser.close()


# Mantener la posicion de SUBIR_A_DRIVE accesible (utilizado por lógica futura de Drive)
DRIVE_CARPETA = "DIAN"


# ---------------------------------------------------------------------------
# Punto de entrada alternativo (modo consola) — útil para pruebas rápidas
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Consulta DIAN individual (web runner)")
    parser.add_argument("--tipo", default="Cédula de Ciudadanía")
    parser.add_argument("--cedula", required=True)
    parser.add_argument("--clave", required=True)
    parser.add_argument("--job-dir", default="./.tmp_job")
    args = parser.parse_args()

    job = Path(args.job_dir)

    async def _main():
        def cb(msg, done):
            print(("[DONE] " if done else "[    ] ") + msg)
        runner = DianRunner(job_dir=job, progreso=cb)
        try:
            final = await runner.consulta_individual(args.tipo, args.cedula, args.clave)
            print(f"[resultado] {final}")
        except Exception as exc:
            print(f"[error] {type(exc).__name__}: {exc}", file=sys.stderr)
            sys.exit(1)

    asyncio.run(_main())