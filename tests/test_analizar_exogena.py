"""Pruebas del análisis de renta (analizar_exogena) con reportes XLSX de muestra."""

import pytest
from openpyxl import Workbook

from app import comun

INGRESOS_UMBRAL = 1_400 * comun.UVT_2025          # >= supera
PATRIMONIO_UMBRAL = 4_500 * comun.UVT_2025        # > supera


def _reporte(path, nombre="JUAN PEREZ", ingresos=None, patrimonio=None):
    """Genera un XLSX con el formato que espera analizar_exogena."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Nombres / Razón social"
    ws["B1"] = nombre
    fila = 2
    if ingresos is not None:
        ws.cell(row=fila, column=5, value="Tope Ingresos brutos")
        ws.cell(row=fila, column=6, value=ingresos)
        fila += 1
    if patrimonio is not None:
        ws.cell(row=fila, column=5, value="Tope Patrimonio bruto")
        ws.cell(row=fila, column=6, value=patrimonio)
        fila += 1
    wb.save(path)
    return path


def test_excede_por_ingresos(runner, tmp_path):
    ruta = _reporte(tmp_path / "exo.xlsx", ingresos=INGRESOS_UMBRAL)
    res = runner.analizar_exogena(ruta)
    assert res["nombre_cliente"] == "JUAN PEREZ"
    assert res["declara_renta"] == "Sí"
    assert "Ingresos brutos" in res["razones"]


def test_excede_por_patrimonio(runner, tmp_path):
    ruta = _reporte(tmp_path / "exo.xlsx", patrimonio=PATRIMONIO_UMBRAL + 1)
    res = runner.analizar_exogena(ruta)
    assert res["declara_renta"] == "Sí"
    assert "Patrimonio bruto" in res["razones"]


def test_no_excede(runner, tmp_path):
    ruta = _reporte(tmp_path / "exo.xlsx",
                    ingresos=INGRESOS_UMBRAL - 1,
                    patrimonio=PATRIMONIO_UMBRAL - 1)
    res = runner.analizar_exogena(ruta)
    assert res["declara_renta"] == "No"
    assert "NO está obligada" in res["razones"]


def test_topes_estructurados(runner, tmp_path):
    """El análisis expone el cotejo de topes en forma estructurada para el panel."""
    ruta = _reporte(tmp_path / "topes.xlsx", ingresos=INGRESOS_UMBRAL)
    res = runner.analizar_exogena(ruta)
    assert isinstance(res["topes"], list) and len(res["topes"]) == 5
    assert set(("desc", "cat", "reportado", "umbral", "excede")) <= set(res["topes"][0])
    ingreso = next(t for t in res["topes"] if t["cat"] == "Ingresos")
    assert ingreso["excede"] is True
    assert ingreso["reportado"] == INGRESOS_UMBRAL
    assert ingreso["umbral"] == INGRESOS_UMBRAL
    assert res["cabecera"]
    assert "IMPORTANTE" in res["razones"]


def test_reportes_vacio_devuelve_desconocido(runner, tmp_path):
    ruta = tmp_path / "vacio.xlsx"
    Workbook().save(ruta)
    res = runner.analizar_exogena(ruta)
    assert res["nombre_cliente"] == "DESCONOCIDO"
    assert res["declara_renta"] == "No"


def test_archivo_corrupto_lanza_error_legible(runner, tmp_path):
    ruta = tmp_path / "corrupto.xlsx"
    ruta.write_bytes(b"esto no es un excel valido")
    with pytest.raises(RuntimeError, match="ilegible o está corrupto"):
        runner.analizar_exogena(ruta)