"""Pruebas del motor de consultas masivas (app/batch.py) con runner simulado."""

import pytest
from openpyxl import Workbook

from app import batch as batch_mod


def _excel(path, filas):
    """Crea un .xlsx con encabezados y filas (listas de dict)."""
    wb = Workbook()
    ws = wb.active
    cab = ["tipo_documento", "numero_documento", "contrasena", "estado"]
    ws.append(cab)
    for f in filas:
        ws.append([f.get("tipo_documento", "Cédula de Ciudadanía"),
                   f.get("numero_documento", ""),
                   f.get("contrasena", ""),
                   f.get("estado", "")])
    wb.save(path)
    return path


def test_cargar_filas_filtra_ok_y_vacias(tmp_path):
    ruta = _excel(tmp_path / "a.xlsx", [
        {"numero_documento": "111", "contrasena": "c1"},
        {"numero_documento": "222", "contrasena": "c2", "estado": "ok"},
        {},  # vacía -> se omite
        {"numero_documento": "333", "contrasena": "c3"},
    ])
    filas, _ = batch_mod.cargar_filas(ruta)
    docs = [f["numero_documento"] for f in filas]
    assert docs == ["111", "333"]
    assert filas[0]["fila_excel"] == 2  # primera fila de datos (1-indexed en Excel)


def test_cargar_filas_error_si_faltan_columnas(tmp_path):
    ruta = tmp_path / "b.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["solo", "una"])
    wb.save(ruta)
    with pytest.raises(ValueError, match="tipo_documento"):
        batch_mod.cargar_filas(ruta)


def test_guardar_estado_escribe_columna(tmp_path):
    ruta = _excel(tmp_path / "c.xlsx", [{"numero_documento": "111", "contrasena": "c1"}])
    batch_mod.guardar_estado(ruta, 2, "ok")
    filas, _ = batch_mod.cargar_filas(ruta)
    assert filas == []  # la fila ok ya no se reprocesa


def test_ejecutar_batch_actualiza_estados(tmp_path, monkeypatch):
    """Flujo completo con DianRunner simulado que genera un .xls por fila."""
    entrada = _excel(tmp_path / "batch.xlsx", [
        {"numero_documento": "111", "contrasena": "c1"},
        {"numero_documento": "222", "contrasena": "c2"},  # fallará credenciales
    ])

    class RunnerStub:
        def __init__(self, job_dir, progreso=None):
            self.job_dir = job_dir
            self.ultima_fecha_vencimiento = "2026-10-16"
            (job_dir / "clientes").mkdir(parents=True, exist_ok=True)

        async def consulta_individual(self, tipo, numero, contrasena):
            if numero == "222":
                raise RuntimeError("Credenciales rechazadas o mensaje de error en la página.")
            final = self.job_dir / "clientes" / f"{numero}.xls"
            final.write_bytes(b"x")
            return final

    monkeypatch.setattr(batch_mod, "DianRunner", RunnerStub)
    eventos = []
    resumen = {}

    async def _main():
        nonlocal resumen
        resumen = await batch_mod.ejecutar_batch(
            tmp_path / "job", entrada,
            lambda g, m, d=None: eventos.append((g, m)),
        )

    import asyncio
    asyncio.run(_main())

    assert resumen["total"] == 2
    assert resumen["ok"] == 1
    assert resumen["error_credenciales"] == 1
    assert resumen["generados"][0].endswith("111.xls")

    # El Excel de entrada quedó actualizado: la fila ok ya no se reprocesa
    filas, _ = batch_mod.cargar_filas(entrada)
    docs = [f["numero_documento"] for f in filas]
    assert docs == ["222"]  # la error_credenciales queda pendiente de reintento


def test_generar_plantilla_columnas_sin_fila(tmp_path):
    ruta = tmp_path / "plantilla.xlsx"
    batch_mod.generar_plantilla(ruta)
    from openpyxl import load_workbook
    wb = load_workbook(ruta)
    ws = wb.active
    cab = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert cab == ["tipo_documento", "numero_documento", "contrasena",
                   "fecha_vencimiento", "estado"]
    # Sin fila de ejemplo: solo debe existir la fila 1 (encabezados)
    assert ws.max_row == 1


def test_guardar_fecha_vencimiento_escribe_columna(tmp_path):
    ruta = _excel(tmp_path / "d.xlsx", [{"numero_documento": "111", "contrasena": "c1"}])
    batch_mod.guardar_fecha_vencimiento(ruta, 2, "2026-10-16")
    from openpyxl import load_workbook
    wb = load_workbook(ruta)
    ws = wb.active
    cab = [(_ or "").strip().lower() for _ in
           next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    assert "fecha_vencimiento" in cab
    col = cab.index("fecha_vencimiento") + 1
    assert ws.cell(row=2, column=col).value == "2026-10-16"
