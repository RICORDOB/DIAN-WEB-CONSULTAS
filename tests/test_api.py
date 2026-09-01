"""Smoke de la API web: flujo de alta→aprobación→bloqueo, revocación en vivo,
dashboard del admin y seguridad básica, usando TestClient."""

import time
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

from app import auth
from app.main import _ratelimit, app


@pytest.fixture(autouse=True)
def _limpiar_ratelimit():
    _ratelimit.clear()


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


def _login(client, usuario, password):
    r = client.post("/api/login", json={"usuario": usuario, "password": password})
    return r


def test_flujo_alta_aprobacion_bloqueo_revocacion(client, db, admin):
    # Alta queda pendiente
    r = client.post("/api/registro", json={"usuario": "pedro", "password": "clave123"})
    assert r.status_code == 200 and r.json()["estado"] == "pendiente"
    assert _login(client, "pedro", "clave123").status_code == 401  # pendiente

    # Admin aprueba
    admin_login = _login(client, admin, "admin123")
    assert admin_login.status_code == 200
    cookie = admin_login.cookies.get("sesion")
    r = client.get("/api/admin/pendientes", cookies={"sesion": "token-invalido"})
    assert r.status_code == 403
    r = client.get("/api/admin/pendientes", cookies={"sesion": cookie})
    assert r.status_code == 200
    assert any(u["usuario"] == "pedro" for u in r.json()["pendientes"])
    r = client.post("/api/admin/decidir", json={"usuario": "pedro", "aprobar": True},
                    cookies={"sesion": cookie})
    assert r.status_code == 200

    # Usuario ya puede entrar
    pedro = _login(client, "pedro", "clave123")
    assert pedro.status_code == 200
    pedro_cookie = pedro.cookies.get("sesion")

    # Bloqueo: derriba la sesión en la siguiente petición y prohíbe el login
    r = client.post("/api/admin/bloquear", json={"usuario": "pedro", "bloquear": True},
                    cookies={"sesion": cookie})
    assert r.status_code == 200 and r.json()["estado"] == "bloqueado"
    r = client.get("/api/me", cookies={"sesion": pedro_cookie})
    assert r.json()["autenticado"] is False  # revocación en vivo
    assert _login(client, "pedro", "clave123").status_code == 401

    # Desbloqueo: vuelve a entrar
    r = client.post("/api/admin/bloquear", json={"usuario": "pedro", "bloquear": False},
                    cookies={"sesion": cookie})
    assert r.status_code == 200
    assert _login(client, "pedro", "clave123").status_code == 200


def test_admin_no_puede_bloquearse(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    r = client.post("/api/admin/bloquear", json={"usuario": admin, "bloquear": True},
                    cookies={"sesion": cookie})
    assert r.status_code == 400


def test_usuarios_comunes_no_acceden_al_panel(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    client.post("/api/registro", json={"usuario": "laura", "password": "clave123"})
    client.post("/api/admin/decidir", json={"usuario": "laura", "aprobar": True},
                cookies={"sesion": cookie})
    laura = _login(client, "laura", "clave123").cookies.get("sesion")
    assert client.get("/api/admin/estadisticas",
                      cookies={"sesion": laura}).status_code == 403
    assert client.get("/api/admin/consultas",
                      cookies={"sesion": laura}).status_code == 403


def test_estadisticas_y_historial(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    auth.registrar_consulta("job-abc", "pedro", "Cédula de Ciudadanía")
    auth.actualizar_consulta("job-abc", auth.ESTADO_DONE, resultado="PEDRO.xls")

    r = client.get("/api/admin/estadisticas", cookies={"sesion": cookie})
    assert r.status_code == 200
    data = r.json()
    assert data["consultas"]["total"] == 1
    assert data["consultas"]["done"] == 1
    assert data["por_dia"][-1]["total"] >= 1

    r = client.get("/api/admin/consultas", cookies={"sesion": cookie})
    assert r.status_code == 200
    filas = r.json()["consultas"]
    assert len(filas) == 1 and filas[0]["usuario"] == "pedro"
    assert "numero_documento" not in filas[0]


def test_consulta_sin_sesion_rechazada(client, db):
    r = client.post("/api/consulta", json={
        "tipo_documento": "Cédula de Ciudadanía",
        "numero_documento": "12345678",
        "contrasena": "x",
    })
    assert r.status_code == 401


def test_admin_elimina_usuario_y_su_historial(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")

    # Alta y aprobación
    client.post("/api/registro", json={"usuario": "ana", "password": "clave123"})
    assert client.post("/api/admin/decidir", json={"usuario": "ana", "aprobar": True},
                       cookies={"sesion": cookie}).status_code == 200

    # Ana entra y genera una consulta persistida
    ana_cookie = _login(client, "ana", "clave123").cookies.get("sesion")
    assert client.get("/api/me", cookies={"sesion": ana_cookie}).json()["autenticado"] is True
    auth.registrar_consulta("job-ana", "ana", "Cédula de Ciudadanía")
    assert auth.listar_consultas(usuario="ana")  # hay historial

    # Eliminación
    r = client.post("/api/admin/eliminar", json={"usuario": "ana"},
                    cookies={"sesion": cookie})
    assert r.status_code == 200 and r.json()["usuario"] == "ana"

    # Cuenta y consultas purgadas; la sesión de Ana queda revocada en vivo
    assert auth.estado_usuario("ana") is None
    assert auth.listar_consultas(usuario="ana") == []
    assert client.get("/api/me", cookies={"sesion": ana_cookie}).json()["autenticado"] is False
    assert _login(client, "ana", "clave123").status_code == 401

    # Ya no figura en la lista ni en las estadísticas
    pendientes = client.get("/api/admin/pendientes", cookies={"sesion": cookie}).json()
    assert not any(u["usuario"] == "ana" for u in pendientes["pendientes"])
    stats = client.get("/api/admin/estadisticas", cookies={"sesion": cookie}).json()
    assert stats["usuarios"]["total"] == 1  # solo el admin


def test_admin_no_puede_eliminarse_ni_eliminar_un_admin(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    r = client.post("/api/admin/eliminar", json={"usuario": admin},
                    cookies={"sesion": cookie})
    assert r.status_code == 400
    r = client.post("/api/admin/eliminar", json={"usuario": "inexistente"},
                    cookies={"sesion": cookie})
    assert r.status_code == 400


def test_admin_activa_y_desactiva_acceso_contador(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")

    # Alta y aprobación de un usuario común
    client.post("/api/registro", json={"usuario": "contab", "password": "clave123"})
    client.post("/api/admin/decidir", json={"usuario": "contab", "aprobar": True},
                cookies={"sesion": cookie})

    # Sin acceso: /contadores redirige a /panel (client limpio para no mezclar jar del admin)
    c = TestClient(app, follow_redirects=False)
    c_cookie = _login(c, "contab", "clave123").cookies.get("sesion")
    r = c.get("/contadores")
    assert r.status_code == 303 and "/panel" in r.headers["location"]
    assert c.get("/api/me").json()["acceso_contador"] is False

    # Admin activa el acceso
    r = client.post("/api/admin/contador", json={"usuario": "contab", "activar": True},
                    cookies={"sesion": cookie})
    assert r.status_code == 200 and r.json()["acceso_contador"] is True

    # Ahora sí ve el panel de contadores
    r = c.get("/contadores")
    assert r.status_code == 200
    assert c.get("/api/me").json()["acceso_contador"] is True

    # Desactiva
    r = client.post("/api/admin/contador", json={"usuario": "contab", "activar": False},
                    cookies={"sesion": cookie})
    assert r.status_code == 200 and r.json()["acceso_contador"] is False
    r = c.get("/contadores")
    assert r.status_code == 303

    # Un usuario común no puede activar a otros
    assert c.post("/api/admin/contador", json={"usuario": "contab", "activar": True}
                  ).status_code == 403


def test_plantilla_masiva_requiere_acceso_contador(client, db, admin):
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    client.post("/api/registro", json={"usuario": "contab2", "password": "clave123"})
    client.post("/api/admin/decidir", json={"usuario": "contab2", "aprobar": True},
                cookies={"sesion": cookie})
    c = TestClient(app, follow_redirects=False)
    c_cookie = _login(c, "contab2", "clave123").cookies.get("sesion")

    # Sin acceso: 403
    r = c.get("/api/masiva/plantilla")
    assert r.status_code == 403

    # Con acceso: 200 y descarga plantilla .xlsx con las columnas esperadas
    r = client.post("/api/admin/contador", json={"usuario": "contab2", "activar": True},
                    cookies={"sesion": cookie})
    assert r.status_code == 200
    r = c.get("/api/masiva/plantilla")
    assert r.status_code == 200
    assert r.headers["content-disposition"].startswith("attachment")
    assert ".xlsx" in r.headers["content-disposition"]

    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    cab = [_ or "" for _ in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    assert cab == ["tipo_documento", "numero_documento", "contrasena",
                   "fecha_vencimiento", "estado"]
    assert ws.max_row == 1  # sin fila de ejemplo


def test_headers_de_seguridad(client, db):
    r = client.get("/")
    assert r.headers.get("x-frame-options") == "DENY"
    assert r.headers.get("x-content-type-options") == "nosniff"
    assert r.headers.get("referrer-policy") == "same-origin"


def test_pagina_privacidad(client, db):
    r = client.get("/privacidad")
    assert r.status_code == 200
    assert "Política de Privacidad" in r.text
    assert "Ricardo Córdoba y Ana María Rozo" in r.text or "Ricardo Cordoba" in r.text
    # El pie conserva la línea de desarrollo y añade la de derechos
    assert "Todos los derechos reservados" in r.text


def test_consulta_expone_resultado_al_panel(client, db, admin, monkeypatch, tmp_path):
    """Un job completado expone el análisis estructurado vía /api/job/{id}."""
    from app import main as mainmod

    class RunnerStub:
        def __init__(self, job_dir, progreso=None):
            self.ultimo_analisis = {
                "nombre_cliente": "ANA",
                "declara_renta": "No",
                "cabecera": "La persona NO está obligada a declarar renta.",
                "nota": "",
                "topes": [{
                    "desc": "Ingresos brutos", "cat": "Ingresos",
                    "reportado": 100.0, "umbral": 200.0, "excede": False,
                }],
            }

        async def consulta_individual(self, *args, **kwargs):
            ruta = tmp_path / "ANA.xls"
            ruta.write_bytes(b"x")
            return ruta

    monkeypatch.setattr(mainmod, "DianRunner", RunnerStub)
    cookie = _login(client, admin, "admin123").cookies.get("sesion")
    r = client.post("/api/consulta", json={
        "tipo_documento": "Cédula de Ciudadanía",
        "numero_documento": "12345678",
        "contrasena": "secreta",
    }, cookies={"sesion": cookie})
    assert r.status_code == 200
    job_id = r.json()["job_id"]

    data = {}
    for _ in range(60):
        data = client.get("/api/job/" + job_id, cookies={"sesion": cookie}).json()
        if data["estado"] in ("done", "error"):
            break
        time.sleep(0.05)

    assert data["estado"] == "done"
    assert data["final"].endswith("ANA.xls")
    assert data["resultado"]["declara_renta"] == "No"
    assert data["resultado"]["nombre_cliente"] == "ANA"
    assert data["resultado"]["topes"][0]["excede"] is False