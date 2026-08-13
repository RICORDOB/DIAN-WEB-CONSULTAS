"""
DIAN - Automatización de ingreso de usuarios registrados (Fase 1: Login)
-----------------------------------------------------------------------
Lee credenciales desde un Excel y automatiza el login en el portal
transaccional de la DIAN (muisca.dian.gov.co).

Stack: Python + Playwright + openpyxl.

Flujo por cliente (Fases 1-3 + Facturación Electrónica):
  - cargar_credenciales(): solo lectura de datos (desacoplada de la UI)
  - iniciar_sesion(): login en el portal MUISCA (reutilizable por fila)
  - consultar_exogena(): descarga el reporte de información exógena
  - consultar_facturacion_electronica(): descarga el reporte FE (o None)
  - analizar_exogena(): determina si debe declarar renta (Art. 592 E.T.)
  - armar_libro_cliente(): une todo en un XLSX de 3 hojas renombrado
    con el nombre del cliente (Hoja1=Exógena, Hoja2=Renta, Hoja3=FE)
  - main(): orquesta navegador + bucle de filas + logs

Uso:
  python3 dian_login.py                 # login en modo visible (headed)
  python3 dian_login.py --headless      # sin ventana (para validar/CI)
  python3 dian_login.py --inspeccionar  # solo navega y vuelca el HTML
                                         # + screenshot para detectar selectores
"""

import argparse
import asyncio
import re
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from playwright.async_api import async_playwright

# ---------------------------------------------------------------------------
# CONFIGURACIÓN
# ---------------------------------------------------------------------------
URL_LOGIN = "https://muisca.dian.gov.co/WebArquitectura/DefLogin.faces"
EXCEL_PATH = Path(__file__).parent / "clientes_dian.xlsx"
EXCEL_INDIVIDUAL_PATH = Path(__file__).parent / "cliente_individual.xlsx"
DOWNLOAD_DIR = Path(__file__).parent / "descargas"
CLIENTES_DIR = Path(__file__).parent / "clientes"

# Tiempos y reintentos
TIMEOUT_LOGIN = 15_000
TIMEOUT_DESCARGA = 30_000
REINTENTOS = 2

# Log de ejecución (consola + archivo en descargas/)
LOG_PATH = DOWNLOAD_DIR / f"log_dian_{datetime.now():%Y%m%d_%H%M%S}.txt"


def loguear(mensaje: str) -> None:
    """Escribe a consola y al archivo de log de la ejecución."""
    ts = datetime.now().strftime("%H:%M:%S")
    linea = f"[{ts}] {mensaje}"
    print(linea)
    LOG_PATH.parent.mkdir(exist_ok=True)
    with open(LOG_PATH, "a", encoding="utf-8") as fh:
        fh.write(linea + "\n")

# Fase 2: año gravable a consultar en "Consultar información Exógena"
ANIO_EXOGENO = "2025"

# Fase 3: determinación de obligación de declarar renta (AG 2025, se presenta en 2026)
# Se usa la UVT 2025 (no la de 2026, que solo aplica a sanciones).
UVT_2025 = 49_799
RESOLUCION_UVT = "Res. DIAN 000193 de 2024"
NORMA_TOPES = "Art. 592 E.T. y Dec. 1625-2016 (art. 1.6.1.13.2.7)"

# Calendario DIAN 2026 — Vencimiento declaración renta personas naturales AG 2025
# Res. DIAN 000238 de 2025. Fechas según los dos últimos dígitos del NIT.
CALENDARIO_RENTA_2026 = {
    (1, 2): date(2026, 8, 12),   (3, 4): date(2026, 8, 13),
    (5, 6): date(2026, 8, 14),   (7, 8): date(2026, 8, 18),
    (9, 10): date(2026, 8, 19),  (11, 12): date(2026, 8, 20),
    (13, 14): date(2026, 8, 21), (15, 16): date(2026, 8, 24),
    (17, 18): date(2026, 8, 25), (19, 20): date(2026, 8, 26),
    (21, 22): date(2026, 8, 27), (23, 24): date(2026, 8, 28),
    (25, 26): date(2026, 8, 31), (27, 28): date(2026, 9, 1),
    (29, 30): date(2026, 9, 2),  (31, 32): date(2026, 9, 3),
    (33, 34): date(2026, 9, 4),  (35, 36): date(2026, 9, 7),
    (37, 38): date(2026, 9, 8),  (39, 40): date(2026, 9, 9),
    (41, 42): date(2026, 9, 10), (43, 44): date(2026, 9, 11),
    (45, 46): date(2026, 9, 14), (47, 48): date(2026, 9, 15),
    (49, 50): date(2026, 9, 16), (51, 52): date(2026, 9, 17),
    (53, 54): date(2026, 9, 18), (55, 56): date(2026, 9, 21),
    (57, 58): date(2026, 9, 22), (59, 60): date(2026, 9, 23),
    (61, 62): date(2026, 9, 24), (63, 64): date(2026, 9, 25),
    (65, 66): date(2026, 9, 28), (67, 68): date(2026, 10, 1),
    (69, 70): date(2026, 10, 2), (71, 72): date(2026, 10, 5),
    (73, 74): date(2026, 10, 6), (75, 76): date(2026, 10, 7),
    (77, 78): date(2026, 10, 8), (79, 80): date(2026, 10, 9),
    (81, 82): date(2026, 10, 13), (83, 84): date(2026, 10, 14),
    (85, 86): date(2026, 10, 15), (87, 88): date(2026, 10, 16),
    (89, 90): date(2026, 10, 19), (91, 92): date(2026, 10, 20),
    (93, 94): date(2026, 10, 21), (95, 96): date(2026, 10, 22),
    (97, 98): date(2026, 10, 23), (99, 0): date(2026, 10, 26),
}


def _fecha_vencimiento_renta(numero_documento: str) -> date:
    """Calcula la fecha de vencimiento de renta AG 2025 según los dos últimos
    dígitos del NIT/documento usando el calendario DIAN 2026."""
    ultimos_dos = int(numero_documento.strip()[-2:])
    for (a, b), fecha in CALENDARIO_RENTA_2026.items():
        if ultimos_dos == a or ultimos_dos == b:
            return fecha
    msg = f"No hay fecha en el calendario para los dígitos '{ultimos_dos:02d}'"
    raise ValueError(msg)

# Cada tope: (categoría en el reporte, UVT, descripción legal, operador)
TOPES = [
    ("Ingresos",   1_400, "Ingresos brutos",                  ">="),
    ("Patrimonio", 4_500, "Patrimonio bruto",                 ">"),
    ("Consumo TC", 1_400, "Consumos con tarjeta de crédito",  ">="),
    ("Movimiento", 1_400, "Consignaciones bancarias",         ">="),
    ("Compras",    1_400, "Compras y consumos totales",       ">="),
]

# Fase 4: Google Drive
SUBIR_A_DRIVE = True
DRIVE_CARPETA = "DIAN"
DRIVE_CREDENTIALS = Path(__file__).parent / "credentials.json"
DRIVE_TOKEN = Path(__file__).parent / "token.json"

# Mapa de columnas del Excel -> claves internas
COL_TIPO = "tipo_documento"
COL_NRO = "numero_documento"
COL_PASS = "contrasena"
COL_VENC = "fecha_vencimiento"
COL_ESTADO = "estado"

# Texto visible de la opción de tipo de documento (debe coincidir con el Excel)
TEXTO_ACEPTO = "Acepto el tratamiento de los datos personales"


# ---------------------------------------------------------------------------
# LECTURA DE DATOS (desacoplada de la UI -> escalable)
# ---------------------------------------------------------------------------
def cargar_credenciales(ruta: Path) -> list[dict]:
    """Lee el Excel y devuelve una lista de dicts con las credenciales."""
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró el archivo Excel: {ruta}")

    wb = load_workbook(ruta, data_only=True)
    ws = wb.active

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        raise ValueError("El Excel está vacío.")

    encabezados = [str(h).strip() if h is not None else "" for h in filas[0]]
    necesarios = {COL_TIPO, COL_NRO, COL_PASS}
    faltan = necesarios - set(encabezados)
    if faltan:
        raise ValueError(f"Faltan columnas en el Excel: {', '.join(faltan)}")

    idx = {col: encabezados.index(col) for col in necesarios}
    col_venc_idx = encabezados.index(COL_VENC) if COL_VENC in encabezados else None
    col_estado_idx = encabezados.index(COL_ESTADO) if COL_ESTADO in encabezados else None
    credenciales = []
    saltados = 0
    for n, fila in enumerate(filas[1:], start=2):
        tipo = fila[idx[COL_TIPO]]
        nro = fila[idx[COL_NRO]]
        clave = fila[idx[COL_PASS]]
        if tipo is None and nro is None and clave is None:
            continue  # fila vacía
        if nro is None or clave is None:
            loguear(f"[aviso] Fila {n} incompleta, se omite.")
            continue
        estado_actual = str(fila[col_estado_idx]).strip().lower() if col_estado_idx is not None and fila[col_estado_idx] is not None else ""
        if estado_actual == "ok":
            saltados += 1
            continue
        tipo_s = str(tipo).strip()
        nro_s = str(nro).strip()
        if col_venc_idx is not None and fila[col_venc_idx] is not None:
            fecha_venc = str(fila[col_venc_idx]).strip()
        else:
            fecha_venc = _fecha_vencimiento_renta(nro_s).isoformat()
        credenciales.append({
            "fila_excel": n,
            "tipo_documento": tipo_s,
            "numero_documento": nro_s,
            "contrasena": str(clave),
            "fecha_vencimiento": fecha_venc,
        })
    if saltados:
        loguear(f"[info] {saltados} cliente(s) ya en estado 'ok' — omitidos.")
    return credenciales


def _rellenar_fechas_vencimiento(ruta: Path) -> None:
    """Llena la columna fecha_vencimiento para TODAS las filas del Excel
    que tengan número de documento pero no tengan fecha de vencimiento."""
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(values_only=True))]

    if COL_VENC not in encabezados or COL_NRO not in encabezados:
        wb.close()
        return

    col_venc_idx = encabezados.index(COL_VENC) + 1
    col_nro_idx = encabezados.index(COL_NRO) + 1
    actualizados = 0
    for fila in range(2, ws.max_row + 1):
        nro = ws.cell(row=fila, column=col_nro_idx).value
        venc = ws.cell(row=fila, column=col_venc_idx).value
        if nro is not None and (venc is None or str(venc).strip() == ""):
            try:
                fecha = _fecha_vencimiento_renta(str(nro).strip()).isoformat()
                ws.cell(row=fila, column=col_venc_idx, value=fecha)
                actualizados += 1
            except (ValueError, IndexError):
                continue
    if actualizados:
        loguear(f"[info] Fecha de vencimiento asignada a {actualizados} cédula(s) en {ruta.name}.")
    wb.save(ruta)


def _asegurar_columnas(ruta: Path, credenciales: list[dict]) -> None:
    """Agrega las columnas fecha_vencimiento y estado al Excel si no existen."""
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(values_only=True))]

    if COL_VENC not in encabezados:
        col_idx = len(encabezados) + 1
        ws.cell(row=1, column=col_idx, value=COL_VENC)
        for i, cred in enumerate(credenciales, start=2):
            ws.cell(row=i, column=col_idx, value=cred["fecha_vencimiento"])
        loguear(f"[info] Columna '{COL_VENC}' agregada a {ruta.name}.")

    if COL_ESTADO not in encabezados:
        col_idx = len(encabezados) + (2 if COL_VENC not in encabezados else 1)
        ws.cell(row=1, column=col_idx, value=COL_ESTADO)
        loguear(f"[info] Columna '{COL_ESTADO}' agregada a {ruta.name}.")

    wb.save(ruta)


def _actualizar_estado_excel(ruta: Path, fila_excel: int, estado: str) -> None:
    """Escribe el estado en la columna estado de la fila correspondiente."""
    wb = load_workbook(ruta)
    ws = wb.active
    encabezados = [str(h).strip() if h is not None else "" for h in next(ws.iter_rows(values_only=True))]
    if COL_ESTADO not in encabezados:
        return
    col_idx = encabezados.index(COL_ESTADO) + 1
    ws.cell(row=fila_excel, column=col_idx, value=estado)
    wb.save(ruta)


# ---------------------------------------------------------------------------
# LOGIN (una acción por fila -> reutilizable)
# ---------------------------------------------------------------------------
async def iniciar_sesion(page, creds: dict) -> str:
    """
    Realiza el login con un conjunto de credenciales en el portal MUISCA.
    Devuelve un estado: "ok", "error_credenciales" o "desconocido".
    """
    await page.goto(URL_LOGIN, wait_until="networkidle")

    # 0) Asegurar "A nombre propio" (persona natural). Es el default, pero
    #    lo seleccionamos explícitamente por robustez.
    await page.get_by_role("button", name="A nombre propio", exact=True).click()

    # 1) Tipo de documento (mat-select de Angular Material).
    #    El texto de las opciones es insensible a mayúsculas/minúsculas.
    await page.locator("mat-select[name='tipoDocumento']").click()
    await page.get_by_role(
        "option", name=creds["tipo_documento"], exact=False
    ).first.click()

    # 2) Número de documento (se habilita tras elegir el tipo)
    await page.fill("input[name='numDocumento']", creds["numero_documento"])

    # 3) Contraseña
    await page.fill("input[name='password']", creds["contrasena"])

    # 4) Checkbox "Acepto el tratamiento de los datos personales"
    await page.locator("mat-checkbox[name='aceptaTratamientoDatos']").click()

    # 5) Botón Ingresar (se habilita al completar el formulario)
    await page.get_by_role("button", name="Ingresar", exact=True).click()

    # 6) Evaluar resultado: si el formulario de login desaparece -> entró.
    try:
        await page.wait_for_selector(
            "mat-select[name='tipoDocumento']", state="detached", timeout=10000
        )
        return "ok"
    except Exception:  # noqa: BLE001
        pass

    # ¿mensaje de error visible en la página?
    error_loc = page.locator("text=/inválid|incorrect|no coincid|error/i")
    if await error_loc.count():
        return "error_credenciales"
    return "desconocido"


async def _intentar_login(page, creds: dict) -> str:
    """Intenta iniciar sesión reintentando solo errores transitorios (excepciones).
    Los estados 'error_credenciales'/'desconocido' se devuelven sin reintentar."""
    ultimo = None
    for intento in range(1, REINTENTOS + 1):
        try:
            return await iniciar_sesion(page, creds)
        except Exception as exc:  # noqa: BLE001
            ultimo = exc
            loguear(f"  [login][reintento {intento}/{REINTENTOS}] excepción: "
                    f"{type(exc).__name__}: {exc}")
            try:
                await page.goto(URL_LOGIN, wait_until="networkidle")
            except Exception:  # noqa: BLE001
                pass
            await asyncio.sleep(1)
    raise ultimo


# ---------------------------------------------------------------------------
# FASE 2 (placeholder) -> aquí se conectará la lógica posterior al login
# ---------------------------------------------------------------------------
async def consultar_exogena(page, anio: str, nro_doc: str) -> Path:
    """
    Fase 2: en el dashboard, abre "Consultar información Exógena",
    acepta condiciones, elige el año y descarga el reporte en Excel.
    Devuelve la ruta del archivo descargado (nombrado con el documento).
    """
    await page.wait_for_timeout(1000)
    await _cerrar_modales_dian(page)
    await page.locator("input[id='vistaDashboard:frmDashboard:btnExogena']").click(force=True)
    await page.wait_for_selector(
        "input[id='vistaDashboard:frmDashboard:btnBuscar']",
        state="visible", timeout=15000,
    )

    await _cerrar_modales_dian(page)
    await page.locator("input[id='vistaDashboard:frmDashboard:btnBuscar']").click(force=True)
    await page.wait_for_selector(
        "select[id='vistaDashboard:frmDashboard:anioSel']",
        state="visible", timeout=15000,
    )

    await page.select_option(
        "select[id='vistaDashboard:frmDashboard:anioSel']", value=anio
    )
    await page.wait_for_timeout(2000)

    DOWNLOAD_DIR.mkdir(exist_ok=True)
    destino = DOWNLOAD_DIR / f"informacion_exogena_{anio}_{nro_doc}.xls"
    for intento in range(1, REINTENTOS + 1):
        try:
            await _cerrar_modales_dian(page)
            async with page.expect_download(timeout=TIMEOUT_DESCARGA) as dl_info:
                await page.locator(
                    "input[id='vistaDashboard:frmDashboard:btnExogenaGenerar']"
                ).click(force=True)
                await _cerrar_modales_dian(page)
                download = await dl_info.value
            await download.save_as(destino)
            if destino.exists() and destino.stat().st_size > 0:
                break
            raise RuntimeError("la descarga quedó vacía")
        except Exception as exc:  # noqa: BLE001
            loguear(f"  [exogena][reintento {intento}/{REINTENTOS}] "
                    f"{type(exc).__name__}: {exc}")
            if intento == REINTENTOS:
                raise
            await page.wait_for_timeout(1500)
    return destino


async def consultar_facturacion_electronica(page, anio: str, nro_doc: str) -> Optional[Path]:
    """
    Descarga el reporte de facturación electrónica (mismo flujo que exógena:
    btnFE -> aceptar -> año -> consultar). Devuelve la ruta del archivo o
    None si la DIAN responde "No se encontró información para el año
    seleccionado" o si la página falla (timeout, error de red, etc.).
    En ningún caso lanza excepción: si no se puede descargar, se devuelve
    None para que el flujo continúe y la hoja 3 quede sin datos.
    """
    try:
        await _cerrar_modales_dian(page)
        await page.locator("input[id='vistaDashboard:frmDashboard:btnFE']").click(force=True)
        await page.wait_for_selector(
            "input[id='vistaDashboard:frmDashboard:btnBuscarFE']",
            state="visible", timeout=15000,
        )
        await _cerrar_modales_dian(page)
        await page.locator("input[id='vistaDashboard:frmDashboard:btnBuscarFE']").click(force=True)
        await page.wait_for_selector(
            "select[id='vistaDashboard:frmDashboard:anioSelFE']",
            state="visible", timeout=15000,
        )
        await page.select_option(
            "select[id='vistaDashboard:frmDashboard:anioSelFE']", value=anio
        )
        await page.wait_for_timeout(2000)

        DOWNLOAD_DIR.mkdir(exist_ok=True)
        destino = DOWNLOAD_DIR / f"facturacion_{anio}_{nro_doc}.xls"
        for intento in range(1, REINTENTOS + 1):
            try:
                await _cerrar_modales_dian(page)
                async with page.expect_download(timeout=TIMEOUT_DESCARGA) as dl_info:
                    await page.locator(
                        "input[id='vistaDashboard:frmDashboard:btnFFGenerar']"
                    ).click(force=True)
                    await _cerrar_modales_dian(page)
                    download = await dl_info.value
                await download.save_as(destino)
                if destino.exists() and destino.stat().st_size > 0:
                    return destino
                raise RuntimeError("la descarga quedó vacía")
            except Exception as exc:  # noqa: BLE001
                # Sin datos: la DIAN muestra el recuadro rojo y no hay descarga.
                if await page.get_by_text(
                    "No se encontró información para el año seleccionado", exact=False
                ).count():
                    loguear("  [FE] sin información para el año seleccionado -> "
                            "se omite hoja 3.")
                    return None
                loguear(f"  [FE][reintento {intento}/{REINTENTOS}] "
                        f"{type(exc).__name__}: {exc}")
                if intento == REINTENTOS:
                    break
                await page.wait_for_timeout(1500)
    except Exception as exc:  # noqa: BLE001
        loguear(f"  [FE][aviso] No se pudo descargar la facturación electrónica "
                f"({type(exc).__name__}: {exc}) -> se omite hoja 3.")
        return None
    return None


def _copiar_valores(origen_ws, destino_ws) -> None:
    for row in origen_ws.iter_rows():
        for cell in row:
            destino_ws[cell.coordinate].value = cell.value


def _nombre_seguro(nombre: str) -> str:
    """Elimina caracteres no válidos en nombres de archivo."""
    return re.sub(r'[\\/*?:<>|"]', "", nombre).strip()


def armar_libro_cliente(exogena_path: Path, analisis: dict,
                        facturacion_path: Optional[Path]) -> Path:
    """
    Toma el XLSX de exógena descargado y le agrega:
      Hoja 2 -> Reporte Renta (Sí/No + razones)
      Hoja 3 -> Facturación Electrónica (reporte o nota si no hay datos)
    y lo guarda renombrado con el nombre del cliente.
    Nota: openpyxl rechaza la extensión .xls aunque el contenido sea xlsx,
    por eso copiamos a un temporal .xlsx para abrirlo.
    """
    nombre = _nombre_seguro(analisis["nombre_cliente"])
    tmp_exo = DOWNLOAD_DIR / f".tmp_exo_{nombre}.xlsx"
    shutil.copy(exogena_path, tmp_exo)
    wb = load_workbook(str(tmp_exo))
    wb.worksheets[0].title = "Información Exógena"  # Hoja 1

    # Hoja 2: Renta
    ws2 = wb.create_sheet("Renta")
    ws2["A1"] = "Nombre del cliente"
    ws2["B1"] = analisis["nombre_cliente"]
    ws2["A2"] = "¿Declara renta?"
    ws2["B2"] = analisis["declara_renta"]
    ws2["A3"] = "Vence (Renta AG 2025)"
    ws2["B3"] = analisis["fecha_vencimiento"]
    ws2["A5"] = "Razones:"
    ws2["A6"] = analisis["razones"]
    ws2["A6"].alignment = Alignment(wrap_text=True, vertical="top")

    # Hoja 3: Facturación Electrónica
    ws3 = wb.create_sheet("Facturación Electrónica")
    if facturacion_path and facturacion_path.exists():
        tmp_fe = DOWNLOAD_DIR / f".tmp_fe_{nombre}.xlsx"
        shutil.copy(facturacion_path, tmp_fe)
        _copiar_valores(load_workbook(str(tmp_fe)).active, ws3)
        tmp_fe.unlink()
    else:
        ws3["A1"] = "No se encontró información para el año seleccionado."

    CLIENTES_DIR.mkdir(exist_ok=True)
    final = CLIENTES_DIR / f"{nombre}.xls"
    wb.save(str(final))
    tmp_exo.unlink()
    return final


def analizar_exogena(ruta: Path) -> dict:
    """
    Fase 3: lee el XLSX de información exógena y determina si la persona
    está obligada a declarar renta (AG 2025) según los topes del Art. 592 E.T.
    Devuelve {"nombre_cliente", "declara_renta", "razones"}.
    """
    df = pd.read_excel(ruta, header=None, engine="openpyxl")

    # --- Nombre del cliente ---
    nombre = None
    for r in range(df.shape[0]):
        for c in range(df.shape[1]):
            cell = df.iloc[r, c]
            if isinstance(cell, str) and "Nombres / Razón social" in cell:
                for cc in range(df.shape[1]):
                    val = df.iloc[r, cc]
                    if isinstance(val, str) and "Nombres" not in val and val.strip():
                        nombre = val.strip()
                        break
                if nombre:
                    break
        if nombre:
            break

    # --- Valores de los 5 topes (filas donde la col 4 empieza con "Tope") ---
    valores = {}
    for r in range(df.shape[0]):
        if df.shape[1] <= 4:
            break
        v0 = df.iloc[r, 4]
        if isinstance(v0, str) and v0.strip().lower().startswith("tope"):
            label = v0.strip()
            raw = df.iloc[r, 5] if df.shape[1] > 5 else None
            try:
                num = float(raw)
            except (TypeError, ValueError):
                num = 0.0
            cat = None
            for key in ("Ingresos", "Patrimonio", "Consumo TC",
                        "Movimiento", "Compras"):
                if key.lower() in label.lower():
                    cat = key
                    break
            if cat is None:
                cat = label
            valores[cat] = num

    # --- Cotejo contra los umbrales legales ---
    lineas = []
    declara = False
    for cat, uvt, desc, op in TOPES:
        valor = valores.get(cat, 0.0)
        umbral = uvt * UVT_2025
        excede = (valor >= umbral) if op == ">=" else (valor > umbral)
        if excede:
            declara = True
        lineas.append(
            f"- {desc} (Tope {cat}): reportado ${valor:,.0f}  |  "
            f"umbral {uvt:,} UVT = ${umbral:,.0f}  ->  "
            f"{'EXCEDE' if excede else 'no excede'}"
        )

    cabecera = (
        "La persona ESTÁ OBLIGADA a presentar declaración de renta (AG 2025) "
        "porque supera al menos uno de los topes del " + NORMA_TOPES + "."
        if declara else
        "La persona NO está obligada a declarar renta (AG 2025) según los "
        "topes del " + NORMA_TOPES + " (no supera ninguno)."
    )
    cuerpo = (
        f"Cotejo con UVT 2025 = ${UVT_2025:,} ({RESOLUCION_UVT}).\n"
        + "\n".join(lineas)
    )
    nota = (
        "\n\nNota: el reporte de exógena de la DIAN incluye consumos con "
        "tarjeta débito en 'Consumo TC'; la norma considera solo tarjeta de "
        "crédito, por lo que ese tope podría estar sobreestimado."
        if "Consumo TC" in valores else ""
    )
    razones = cabecera + "\n" + cuerpo + nota

    return {
        "nombre_cliente": nombre or "DESCONOCIDO",
        "declara_renta": "Sí" if declara else "No",
        "razones": razones,
    }


# ---------------------------------------------------------------------------
# FASE 4: Google Drive
# ---------------------------------------------------------------------------
def autenticar_drive():
    """Autentica con Google Drive (OAuth) y devuelve el servicio. Requiere
    credentials.json y genera token.json en el primer uso."""
    from google.auth.transport.requests import Request
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from googleapiclient.discovery import build

    SCOPES = ["https://www.googleapis.com/auth/drive.file"]
    creds = None
    if DRIVE_TOKEN.exists():
        creds = Credentials.from_authorized_user_file(str(DRIVE_TOKEN), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not DRIVE_CREDENTIALS.exists():
                raise FileNotFoundError(
                    f"Falta {DRIVE_CREDENTIALS}. Crea credenciales OAuth de la Drive API."
                )
            flow = InstalledAppFlow.from_client_secrets_file(
                str(DRIVE_CREDENTIALS), SCOPES
            )
            creds = flow.run_local_server(port=0)
        DRIVE_TOKEN.write_text(creds.to_json())
    return build("drive", "v3", credentials=creds)


def _id_carpeta_drive(service, nombre: str) -> str:
    q = (f"mimeType='application/vnd.google-apps.folder' and "
         f"name='{nombre}' and trashed=false")
    res = service.files().list(q=q, spaces="drive",
                               fields="files(id, name)").execute()
    items = res.get("files", [])
    if items:
        return items[0]["id"]
    meta = {"name": nombre, "mimeType": "application/vnd.google-apps.folder"}
    return service.files().create(body=meta, fields="id").execute()["id"]


def subir_a_drive(service, ruta: Path, nombre_carpeta: str = DRIVE_CARPETA) -> None:
    """Sube 'ruta' a la carpeta 'nombre_carpeta' de Google Drive (la crea si no existe)."""
    from googleapiclient.http import MediaFileUpload

    carpeta_id = _id_carpeta_drive(service, nombre_carpeta)
    file_meta = {"name": ruta.name, "parents": [carpeta_id]}
    media = MediaFileUpload(str(ruta), resumable=True)
    service.files().create(body=file_meta, media_body=media, fields="id").execute()


async def _cerrar_modales_dian(page) -> None:
    """Busca y cierra modales visibles de la DIAN (contraseña por
    vencer/vencida) sin tocar las máscaras de RichFaces."""
    for _ in range(40):
        try:
            btn = page.locator("text=Cerrar").first
            if await btn.is_visible(timeout=500):
                await btn.click()
                await page.wait_for_timeout(500)
                continue
        except Exception:
            pass
        try:
            btn = page.locator("text=Aceptar").first
            if await btn.is_visible(timeout=500):
                await btn.click()
                await page.wait_for_timeout(500)
                continue
        except Exception:
            pass
        return


async def post_login(page, creds: dict) -> Path:
    """
    Fase 2 + 3 + FE: tras login, descarga exógena y facturación electrónica,
    determina la obligación de renta y arma el libro de 3 hojas renombrado
    con el nombre del cliente. Devuelve la ruta del archivo final.
    """
    await page.wait_for_timeout(2000)
    await _cerrar_modales_dian(page)
    loguear("  [Fase2] Consultando información exógena...")
    exogena = await consultar_exogena(page, ANIO_EXOGENO, creds["numero_documento"])
    loguear(f"  [Fase2] Exógena descargada: {exogena}")

    loguear("  [Fase3] Analizando obligación de declarar renta...")
    analisis = analizar_exogena(exogena)
    analisis["fecha_vencimiento"] = creds.get("fecha_vencimiento", "")
    loguear(f"  [Fase3] {analisis['nombre_cliente']} -> declara renta: "
            f"{analisis['declara_renta']} | vence: {analisis['fecha_vencimiento']}")

    loguear("  [Fase2] Consultando facturación electrónica...")
    facturacion = await consultar_facturacion_electronica(
        page, ANIO_EXOGENO, creds["numero_documento"]
    )

    final = armar_libro_cliente(exogena, analisis, facturacion)
    loguear(f"  [ok] Archivo cliente generado: {final}")
    return final


# ---------------------------------------------------------------------------
# ORQUESTACIÓN
# ---------------------------------------------------------------------------
async def main(headless: bool, inspeccionar: bool, individual: bool = False):
    ruta_excel = EXCEL_INDIVIDUAL_PATH if individual else EXCEL_PATH

    if not ruta_excel.exists():
        raise FileNotFoundError(f"No se encontró el archivo: {ruta_excel}")

    _asegurar_columnas(ruta_excel, [])
    _rellenar_fechas_vencimiento(ruta_excel)
    credenciales = cargar_credenciales(ruta_excel)
    loguear(f"[info] {len(credenciales)} cliente(s) por procesar")

    stats = {"ok": 0, "error_credenciales": 0, "desconocido": 0, "excepcion": 0}
    finales = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(
            headless=headless,
            slow_mo=500 if not headless else 0,
        )

        if inspeccionar:
            context = await browser.new_context(accept_downloads=True, locale="es-CO")
            page = await context.new_page()
            await page.goto(URL_LOGIN, wait_until="networkidle")
            html = await page.content()
            Path("inspeccion_login.html").write_text(html, encoding="utf-8")
            await page.screenshot(path="inspeccion_login.png", full_page=True)
            loguear("[inspeccion] HTML -> inspeccion_login.html | captura -> inspeccion_login.png")
            await context.close()
            await browser.close()
            return

        drive_service = None
        for i, creds in enumerate(credenciales, start=1):
            loguear(f"\n[{i}/{len(credenciales)}] Login: {creds['tipo_documento']} {creds['numero_documento']}")
            context = await browser.new_context(accept_downloads=True, locale="es-CO")
            page = await context.new_page()
            estado_final = None
            try:
                estado_login = await _intentar_login(page, creds)
                if estado_login == "ok":
                    loguear("  [ok] Sesión iniciada correctamente.")
                    try:
                        final = await post_login(page, creds)
                        finales.append(final)
                        stats["ok"] += 1
                        if SUBIR_A_DRIVE:
                            try:
                                if drive_service is None:
                                    drive_service = autenticar_drive()
                                subir_a_drive(drive_service, final)
                                loguear(f"  [Drive] Subido a carpeta '{DRIVE_CARPETA}': {final.name}")
                            except Exception as exc:
                                loguear(f"  [Drive][aviso] no se pudo subir: {type(exc).__name__}: {exc}")
                        estado_final = "ok"
                    except Exception as exc:
                        loguear(f"  [excepcion post-login] {type(exc).__name__}: {exc}")
                        await page.screenshot(path=f"excepcion_login_{i}.png")
                        stats["excepcion"] += 1
                        estado_final = "excepcion"
                elif estado_login == "error_credenciales":
                    loguear("  [error] Credenciales rechazadas o mensaje de error en la página.")
                    await page.screenshot(path=f"error_login_{i}.png")
                    stats["error_credenciales"] += 1
                    estado_final = "error_credenciales"
                else:
                    loguear("  [aviso] No se pudo determinar el resultado del login.")
                    await page.screenshot(path=f"duda_login_{i}.png")
                    stats["desconocido"] += 1
                    estado_final = "desconocido"
            except Exception as exc:  # noqa: BLE001
                loguear(f"  [excepcion] {type(exc).__name__}: {exc}")
                await page.screenshot(path=f"excepcion_login_{i}.png")
                stats["excepcion"] += 1
                estado_final = "excepcion"
            finally:
                if estado_final and "fila_excel" in creds:
                    _actualizar_estado_excel(ruta_excel, creds["fila_excel"], estado_final)
                await context.close()

    loguear(
        f"\n=== RESUMEN ===  OK={stats['ok']}  error_credenciales={stats['error_credenciales']}  "
        f"desconocido={stats['desconocido']}  excepcion={stats['excepcion']}"
    )
    if finales:
        loguear(f"Archivos de cliente generados ({len(finales)}):")
        for f in finales:
            loguear(f"  - {f}")
    else:
        loguear("No se generó ningún archivo de cliente.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Login DIAN desde Excel")
    parser.add_argument("--headless", action="store_true", help="Ejecutar sin ventana")
    parser.add_argument("--inspeccionar", action="store_true",
                        help="Solo navegar y volcar HTML + captura para detectar selectores")
    parser.add_argument("--individual", action="store_true",
                        help="Usar cliente_individual.xlsx en lugar de clientes_dian.xlsx")
    args = parser.parse_args()
    try:
        asyncio.run(main(headless=args.headless, inspeccionar=args.inspeccionar,
                         individual=args.individual))
    except Exception as exc:  # noqa: BLE001
        print(f"[fatal] {type(exc).__name__}: {exc}", file=sys.stderr)
        loguear(f"[fatal] {type(exc).__name__}: {exc}")
        sys.exit(1)
